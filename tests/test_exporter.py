import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from public_directory_scraper.exporter import write_records, write_records_excel


class WriteRecordsCsvTest(unittest.TestCase):
    def test_writes_records_to_csv(self):
        records = [
            {"name": "Example Business", "url": "https://example.com"},
            {"name": "Second Business", "url": "https://second.example"},
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "listings.csv"

            count = write_records(records, output_path)

            with output_path.open(encoding="utf-8", newline="") as csv_file:
                rows = list(csv.DictReader(csv_file))

        self.assertEqual(count, 2)
        self.assertEqual(rows, records)

    def test_writes_books_records_to_csv(self):
        records = [
            {
                "title": "A Light in the Attic",
                "price": "£51.77",
                "availability": "In stock",
                "rating": "Three",
                "book_url": "catalogue/a-light-in-the-attic_1000/index.html",
                "image_url": "media/cache/book.jpg",
            }
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "books.csv"

            count = write_records(records, output_path)

            with output_path.open(encoding="utf-8", newline="") as csv_file:
                reader = csv.DictReader(csv_file)
                fieldnames = reader.fieldnames
                rows = list(reader)

        self.assertEqual(count, 1)
        self.assertEqual(
            fieldnames,
            ["title", "price", "availability", "rating", "book_url", "image_url"],
        )
        self.assertEqual(rows, records)

    def test_writes_records_to_excel(self):
        records = [
            {
                "title": "A Light in the Attic",
                "price_gbp": 51.77,
                "availability": "In stock",
            }
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "books.xlsx"

            count = write_records_excel(records, output_path)

            workbook = load_workbook(output_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(count, 1)
        self.assertEqual(
            rows,
            [
                ("title", "price_gbp", "availability"),
                ("A Light in the Attic", 51.77, "In stock"),
            ],
        )
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertTrue(sheet["A1"].font.bold)
        self.assertGreaterEqual(sheet.column_dimensions["A"].width, 20)
        self.assertGreaterEqual(sheet.column_dimensions["B"].width, 12)

    def test_rejects_unknown_output_extension(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "books.txt"

            with self.assertRaisesRegex(ValueError, "output file must end"):
                write_records([], output_path)


if __name__ == "__main__":
    unittest.main()
