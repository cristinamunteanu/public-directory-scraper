import csv
import io
import json
import os
import subprocess
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from public_directory_scraper.__main__ import main
from public_directory_scraper.config import EtlConfig
from public_directory_scraper.fetcher import FetchResult
from public_directory_scraper.pipeline import EtlResult

PROJECT_ROOT = Path(__file__).resolve().parents[1]

BOOKS_HTML = """
<article class="product_pod">
  <p class="star-rating Three"></p>
  <h3>
    <a
      href="catalogue/a-light-in-the-attic_1000/index.html"
      title="A Light in the Attic"
    >Book</a>
  </h3>
  <img src="media/cache/book.jpg">
  <p class="price_color">£51.77</p>
  <p class="instock availability">In stock</p>
</article>
"""

CLEANED_RECORDS = [
    {
        "title": "A Light in the Attic",
        "price_gbp": 51.77,
        "availability": "In stock",
        "rating": 3,
        "book_url": "https://books.toscrape.com/book.html",
        "image_url": "https://books.toscrape.com/book.jpg",
    }
]


class CliEntrypointTest(unittest.TestCase):
    def test_package_runs_as_module(self):
        env = os.environ.copy()
        env["PYTHONPATH"] = str(PROJECT_ROOT / "src")

        result = subprocess.run(
            [sys.executable, "-m", "public_directory_scraper"],
            check=True,
            capture_output=True,
            env=env,
            text=True,
        )

        self.assertEqual(result.stdout.strip(), "public-directory-scraper 0.1.0")
        self.assertEqual(result.stderr, "")

    def test_parse_command_prints_books_json(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "books.html"
            input_path.write_text(BOOKS_HTML, encoding="utf-8")
            stdout = io.StringIO()
            stderr = io.StringIO()

            with redirect_stdout(stdout), redirect_stderr(stderr):
                exit_code = main(["parse", str(input_path)])

        self.assertEqual(exit_code, 0)
        self.assertEqual(
            json.loads(stdout.getvalue()),
            [
                {
                    "title": "A Light in the Attic",
                    "price": "£51.77",
                    "availability": "In stock",
                    "rating": "Three",
                    "book_url": "catalogue/a-light-in-the-attic_1000/index.html",
                    "image_url": "media/cache/book.jpg",
                }
            ],
        )
        self.assertEqual(stderr.getvalue(), "")

    def test_fetch_command_prints_status_and_bytes(self):
        stdout = io.StringIO()
        stderr = io.StringIO()

        with (
            patch("public_directory_scraper.__main__.fetch_url") as fetch_url,
            redirect_stdout(stdout),
            redirect_stderr(stderr),
        ):
            fetch_url.return_value = FetchResult(
                status_code=200,
                reason="OK",
                body=b"hello",
            )

            exit_code = main(["fetch", "https://example.com"])

        self.assertEqual(exit_code, 0)
        self.assertEqual(stdout.getvalue(), "200 OK\nbytes: 5\n")
        self.assertEqual(stderr.getvalue(), "")
        fetch_url.assert_called_once_with(
            "https://example.com",
            timeout=10.0,
            retries=0,
        )

    def test_fetch_command_accepts_timeout_and_retries_options(self):
        with (
            patch("public_directory_scraper.__main__.fetch_url") as fetch_url,
            redirect_stdout(io.StringIO()),
            redirect_stderr(io.StringIO()),
        ):
            fetch_url.return_value = FetchResult(
                status_code=200,
                reason="OK",
                body=b"hello",
            )

            exit_code = main(
                [
                    "fetch",
                    "https://example.com",
                    "--timeout",
                    "2.5",
                    "--retries",
                    "1",
                ]
            )

        self.assertEqual(exit_code, 0)
        fetch_url.assert_called_once_with("https://example.com", timeout=2.5, retries=1)

    def test_fetch_command_rejects_negative_retries(self):
        stdout = io.StringIO()
        stderr = io.StringIO()

        with redirect_stdout(stdout), redirect_stderr(stderr):
            exit_code = main(["fetch", "https://example.com", "--retries", "-1"])

        self.assertEqual(exit_code, 2)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(
            stderr.getvalue().strip(),
            "Error: --retries must be zero or a positive integer",
        )

    def test_fetch_command_rejects_invalid_timeout(self):
        stdout = io.StringIO()
        stderr = io.StringIO()

        with redirect_stdout(stdout), redirect_stderr(stderr):
            exit_code = main(["fetch", "https://example.com", "--timeout", "0"])

        self.assertEqual(exit_code, 2)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(
            stderr.getvalue().strip(),
            "Error: --timeout must be a positive number",
        )

    def test_scrape_command_prints_books_json(self):
        stdout = io.StringIO()
        stderr = io.StringIO()

        with (
            patch("public_directory_scraper.__main__.scrape_pages") as scrape_pages,
            redirect_stdout(stdout),
            redirect_stderr(stderr),
        ):
            scrape_pages.return_value = CLEANED_RECORDS

            exit_code = main(["scrape", "https://books.toscrape.com/"])

        self.assertEqual(exit_code, 0)
        self.assertEqual(json.loads(stdout.getvalue()), CLEANED_RECORDS)
        self.assertEqual(stderr.getvalue(), "")

    def test_scrape_command_writes_books_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "books.csv"
            stdout = io.StringIO()
            stderr = io.StringIO()

            with (
                patch("public_directory_scraper.__main__.scrape_pages") as scrape_pages,
                redirect_stdout(stdout),
                redirect_stderr(stderr),
            ):
                scrape_pages.return_value = CLEANED_RECORDS

                exit_code = main(
                    [
                        "scrape",
                        "https://books.toscrape.com/",
                        "--output",
                        str(output_path),
                    ]
                )

            with output_path.open(encoding="utf-8", newline="") as csv_file:
                reader = csv.DictReader(csv_file)
                fieldnames = reader.fieldnames
                rows = list(reader)

        self.assertEqual(exit_code, 0)
        self.assertEqual(stdout.getvalue().strip(), f"Wrote 1 records to {output_path}")
        self.assertEqual(stderr.getvalue(), "")
        self.assertEqual(
            fieldnames,
            ["title", "price_gbp", "availability", "rating", "book_url", "image_url"],
        )
        self.assertEqual(rows[0]["title"], "A Light in the Attic")
        self.assertEqual(rows[0]["price_gbp"], "51.77")

    def test_scrape_command_writes_books_excel(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "books.xlsx"
            stdout = io.StringIO()
            stderr = io.StringIO()

            with (
                patch("public_directory_scraper.__main__.scrape_pages") as scrape_pages,
                redirect_stdout(stdout),
                redirect_stderr(stderr),
            ):
                scrape_pages.return_value = CLEANED_RECORDS

                exit_code = main(
                    [
                        "scrape",
                        "https://books.toscrape.com/",
                        "--output",
                        str(output_path),
                    ]
                )

            workbook = load_workbook(output_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(exit_code, 0)
        self.assertEqual(stdout.getvalue().strip(), f"Wrote 1 records to {output_path}")
        self.assertEqual(stderr.getvalue(), "")
        self.assertEqual(
            rows[0],
            ("title", "price_gbp", "availability", "rating", "book_url", "image_url"),
        )
        self.assertEqual(rows[1][0], "A Light in the Attic")
        self.assertEqual(rows[1][1], 51.77)

    def test_scrape_command_accepts_runtime_options(self):
        with (
            patch("public_directory_scraper.__main__.scrape_pages") as scrape_pages,
            redirect_stdout(io.StringIO()),
            redirect_stderr(io.StringIO()),
        ):
            scrape_pages.return_value = CLEANED_RECORDS

            exit_code = main(
                [
                    "scrape",
                    "https://books.toscrape.com/",
                    "--pages",
                    "2",
                    "--timeout",
                    "5",
                    "--delay",
                    "1.5",
                    "--retries",
                    "1",
                ]
            )

        self.assertEqual(exit_code, 0)
        scrape_pages.assert_called_once_with(
            "https://books.toscrape.com/",
            max_pages=2,
            timeout=5.0,
            delay=1.5,
            retries=1,
        )

    def test_scrape_command_rejects_negative_delay(self):
        stdout = io.StringIO()
        stderr = io.StringIO()

        with redirect_stdout(stdout), redirect_stderr(stderr):
            exit_code = main(
                ["scrape", "https://books.toscrape.com/", "--delay", "-1"]
            )

        self.assertEqual(exit_code, 2)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(
            stderr.getvalue().strip(),
            "Error: --delay must be zero or a positive number",
        )

    def test_parse_command_writes_books_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "books.html"
            input_path.write_text(BOOKS_HTML, encoding="utf-8")
            output_path = Path(temp_dir) / "books.csv"
            stdout = io.StringIO()
            stderr = io.StringIO()

            with redirect_stdout(stdout), redirect_stderr(stderr):
                exit_code = main(
                    ["parse", str(input_path), "--output", str(output_path)]
                )

            with output_path.open(encoding="utf-8", newline="") as csv_file:
                reader = csv.DictReader(csv_file)
                fieldnames = reader.fieldnames
                rows = list(reader)

        self.assertEqual(exit_code, 0)
        self.assertEqual(stdout.getvalue().strip(), f"Wrote 1 records to {output_path}")
        self.assertEqual(stderr.getvalue(), "")
        self.assertEqual(
            fieldnames,
            ["title", "price", "availability", "rating", "book_url", "image_url"],
        )
        self.assertEqual(rows[0]["title"], "A Light in the Attic")
        self.assertEqual(rows[0]["price"], "£51.77")

    def test_scrape_command_reports_output_write_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "missing" / "books.csv"
            stdout = io.StringIO()
            stderr = io.StringIO()

            with (
                patch("public_directory_scraper.__main__.scrape_pages") as scrape_pages,
                redirect_stdout(stdout),
                redirect_stderr(stderr),
            ):
                scrape_pages.return_value = CLEANED_RECORDS

                exit_code = main(
                    [
                        "scrape",
                        "https://books.toscrape.com/",
                        "--output",
                        str(output_path),
                    ]
                )

        self.assertEqual(exit_code, 1)
        self.assertEqual(stdout.getvalue(), "")
        self.assertTrue(
            stderr.getvalue().startswith(f"Error: could not write {output_path}: ")
        )

    def test_parse_command_reports_output_write_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "books.html"
            input_path.write_text(BOOKS_HTML, encoding="utf-8")
            output_path = Path(temp_dir) / "missing" / "books.csv"
            stdout = io.StringIO()
            stderr = io.StringIO()

            with redirect_stdout(stdout), redirect_stderr(stderr):
                exit_code = main(
                    ["parse", str(input_path), "--output", str(output_path)]
                )

        self.assertEqual(exit_code, 1)
        self.assertEqual(stdout.getvalue(), "")
        self.assertTrue(
            stderr.getvalue().startswith(f"Error: could not write {output_path}: ")
        )

    def test_parse_command_reports_missing_file(self):
        missing_path = Path("/tmp/missing-books-page.html")
        stdout = io.StringIO()
        stderr = io.StringIO()

        with redirect_stdout(stdout), redirect_stderr(stderr):
            exit_code = main(["parse", str(missing_path)])

        self.assertEqual(exit_code, 1)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(
            stderr.getvalue().strip(),
            f"Error: file not found: {missing_path}",
        )

    def test_parse_command_reports_invalid_books_html(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "empty.html"
            input_path.write_text("<article></article>", encoding="utf-8")
            stdout = io.StringIO()
            stderr = io.StringIO()

            with redirect_stdout(stdout), redirect_stderr(stderr):
                exit_code = main(["parse", str(input_path)])

        self.assertEqual(exit_code, 1)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(
            stderr.getvalue().strip(),
            "Error: book listing must include title and book_url",
        )

    def test_etl_command_runs_pipeline(self):
        connection = FakeConnection()
        stdout = io.StringIO()
        stderr = io.StringIO()

        with (
            patch("public_directory_scraper.__main__.load_env_file") as load_env_file,
            patch("public_directory_scraper.__main__.load_config") as load_config,
            patch("public_directory_scraper.__main__.connect") as connect,
            patch("public_directory_scraper.__main__.create_tables") as create_tables,
            patch("public_directory_scraper.__main__.run_books_etl") as run_books_etl,
            patch("public_directory_scraper.__main__.logger") as logger,
            redirect_stdout(stdout),
            redirect_stderr(stderr),
        ):
            load_config.return_value = EtlConfig(
                database_url="postgresql://localhost/public_directory_scraper",
                default_pages=3,
                default_timeout=7,
                default_retries=2,
                default_delay=0.5,
            )
            connect.return_value = connection
            run_books_etl.return_value = EtlResult(
                run_id="run-1",
                raw_count=4,
                cleaned_count=3,
            )

            exit_code = main(
                [
                    "etl",
                    "https://books.toscrape.com/",
                    "--run-id",
                    "run-1",
                ]
            )

        self.assertEqual(exit_code, 0)
        self.assertEqual(
            stdout.getvalue(),
            "Raw records loaded: 4\nCleaned records loaded: 3\n",
        )
        self.assertEqual(stderr.getvalue(), "")
        load_env_file.assert_called_once_with()
        connect.assert_called_once_with(
            "postgresql://localhost/public_directory_scraper"
        )
        create_tables.assert_called_once_with(connection)
        run_books_etl.assert_called_once_with(
            connection,
            "https://books.toscrape.com/",
            run_id="run-1",
            pages=3,
            timeout=7,
            delay=0.5,
            retries=2,
        )
        self.assertEqual(logger.info.call_count, 2)
        logger.info.assert_any_call(
            "Starting ETL run %s for %s "
            "(pages=%s, timeout=%s, delay=%s, retries=%s)",
            "run-1",
            "https://books.toscrape.com/",
            3,
            7,
            0.5,
            2,
        )
        logger.info.assert_any_call(
            "Finished ETL run %s: raw_count=%s cleaned_count=%s",
            "run-1",
            4,
            3,
        )
        self.assertTrue(connection.closed)

    def test_etl_command_accepts_runtime_options(self):
        connection = FakeConnection()

        with (
            patch("public_directory_scraper.__main__.load_env_file"),
            patch("public_directory_scraper.__main__.load_config") as load_config,
            patch("public_directory_scraper.__main__.connect") as connect,
            patch("public_directory_scraper.__main__.create_tables"),
            patch("public_directory_scraper.__main__.run_books_etl") as run_books_etl,
            redirect_stdout(io.StringIO()),
            redirect_stderr(io.StringIO()),
        ):
            load_config.return_value = EtlConfig(
                database_url="postgresql://localhost/public_directory_scraper",
            )
            connect.return_value = connection
            run_books_etl.return_value = EtlResult(
                run_id="run-1",
                raw_count=1,
                cleaned_count=1,
            )

            exit_code = main(
                [
                    "etl",
                    "https://books.toscrape.com/",
                    "--run-id",
                    "run-1",
                    "--pages",
                    "2",
                    "--timeout",
                    "5",
                    "--delay",
                    "1.5",
                    "--retries",
                    "1",
                ]
            )

        self.assertEqual(exit_code, 0)
        self.assertEqual(run_books_etl.call_args.kwargs["pages"], 2)
        self.assertEqual(run_books_etl.call_args.kwargs["timeout"], 5.0)
        self.assertEqual(run_books_etl.call_args.kwargs["delay"], 1.5)
        self.assertEqual(run_books_etl.call_args.kwargs["retries"], 1)

    def test_etl_command_reports_missing_database_url(self):
        stdout = io.StringIO()
        stderr = io.StringIO()

        with (
            patch("public_directory_scraper.__main__.load_env_file"),
            patch("public_directory_scraper.__main__.load_config") as load_config,
            redirect_stdout(stdout),
            redirect_stderr(stderr),
        ):
            load_config.side_effect = ValueError("DATABASE_URL is required")

            exit_code = main(
                [
                    "etl",
                    "https://books.toscrape.com/",
                    "--run-id",
                    "run-1",
                ]
            )

        self.assertEqual(exit_code, 2)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(stderr.getvalue().strip(), "Error: DATABASE_URL is required")

    def test_etl_command_logs_runtime_failure(self):
        connection = FakeConnection()
        stdout = io.StringIO()
        stderr = io.StringIO()

        with (
            patch("public_directory_scraper.__main__.load_env_file"),
            patch("public_directory_scraper.__main__.load_config") as load_config,
            patch("public_directory_scraper.__main__.connect") as connect,
            patch("public_directory_scraper.__main__.create_tables"),
            patch("public_directory_scraper.__main__.run_books_etl") as run_books_etl,
            patch("public_directory_scraper.__main__.logger") as logger,
            redirect_stdout(stdout),
            redirect_stderr(stderr),
        ):
            load_config.return_value = EtlConfig(
                database_url="postgresql://localhost/public_directory_scraper",
            )
            connect.return_value = connection
            run_books_etl.side_effect = OSError("load failed")

            exit_code = main(
                [
                    "etl",
                    "https://books.toscrape.com/",
                    "--run-id",
                    "run-1",
                ]
            )

        self.assertEqual(exit_code, 1)
        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(
            stderr.getvalue().strip(),
            "Error: could not run ETL for https://books.toscrape.com/: load failed",
        )
        logger.exception.assert_called_once_with(
            "ETL run %s failed for %s",
            "run-1",
            "https://books.toscrape.com/",
        )
        self.assertTrue(connection.closed)

    def test_etl_command_does_not_hide_unexpected_errors(self):
        connection = FakeConnection()

        with (
            patch("public_directory_scraper.__main__.load_env_file"),
            patch("public_directory_scraper.__main__.load_config") as load_config,
            patch("public_directory_scraper.__main__.connect") as connect,
            patch("public_directory_scraper.__main__.create_tables"),
            patch("public_directory_scraper.__main__.run_books_etl") as run_books_etl,
            redirect_stdout(io.StringIO()),
            redirect_stderr(io.StringIO()),
        ):
            load_config.return_value = EtlConfig(
                database_url="postgresql://localhost/public_directory_scraper",
            )
            connect.return_value = connection
            run_books_etl.side_effect = RuntimeError("programming mistake")

            with self.assertRaisesRegex(RuntimeError, "programming mistake"):
                main(
                    [
                        "etl",
                        "https://books.toscrape.com/",
                        "--run-id",
                        "run-1",
                    ]
                )

        self.assertTrue(connection.closed)


class FakeConnection:
    def __init__(self):
        self.closed = False

    def close(self):
        self.closed = True


if __name__ == "__main__":
    unittest.main()
